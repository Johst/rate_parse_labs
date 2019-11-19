## META INFO ##
# WHAT IS THIS: A script for generic supplier rate deck update. See readme in repo https://github.com/Johst/RateParse.git
# INPUT: A supplier "new_file" xlsx file with new rates, a "current_file" CSV supplier rate file downloaded from system
# OUTPUT: a new re-named "current_file" CSV for re-upload to system
# Target supplier for development = Deutsche Telekom's rate deck.

## FILE LOGISTICS ##
# Files folder: C:\Users\johastje\Desktop\PythonPlay\Costlists\generic_supplier
# Code module: C:\Users\johastje\Desktop\PythonPlay\Costlists\generic_supplier\rateupdate_supplier_generic.py
# Code execution: C:/Users/johastje/AppData/Local/Programs/Python/Python36/python.exe C:\Users\johastje\Desktop\PythonPlay\Costlists\generic_supplier\rateupdate_supplier_generic.py

## CODING PRIORITY ##
# Frequent update senders: Deutsche Telekom (2-3 times per week), CLX (daily) 
# Most complex format senders: Telenor (monthly)

## FUNCTIONS ##
# (Done) create_dict_from_new(new_cost_file) - creates dict with new rates from supplier file
# (Done) create_dict_from_system(system_current) - creates dict with current rates from system file
# (Done) create_new_system_file(new_rates_dict) - Create new system file for upload with new rates


import openpyxl, pprint, csv
from openpyxl import load_workbook

## Global variables - Specific Supplier Configuration:
supplier_name_pos = 'Deutsche Telekom AG' # Manual if not file pos [col, row] can be set
account_product = '' # # Manual if not file pos [col, row] can be set
validity_date_col = 4 #Column for "valid date" value 
#validity_date_pos = [x, y] # Read from a specific file position if not on every network row
currency_pos = 'EUR' # Manual if not file pos [col, row] can be set
data_start_row = 19 # First row of network data
mccmnc_col = 3 # Column for concatenated mccmnc value 
rate_col = 5 # Column for concatenated rate value 
change_status_col = 6 # Column "Increase", "Decrease", "Unchanged" value
fx_conversion = 100 # Conversion multiplication value for FX is "1", "100" 
                    # or "0.01". I.e if supplier rate is in EUR it has to be 
                    # mulitplied by 100 to become EURO cent as in system.
rate_decimal_type = ',' # If ',' (comma), replace to '.' (full stop) as system file 
                        # requires it. IGNORE coding it for now, seems Python takes care of it.    

def create_dict_from_supplier_file(new_cost_file):
    """ Create a dict with new prices derived from the new cost list from [xxx] supplier with its unique format.
    WHAT: Convert supplier Excel file to a dictinary with mccmnc-rate value pairs
    ARGS: Supplier new costlist file
    RETURN: new MCCMNC:rate dict
    """
    ## Dev status: 
    # 2019-11-14 Done and tested!

    ## Specific Supplier Configuration:
    supplier_name_pos = 'Deutsche Telekom AG' # Manual if not file pos [col, row] can be set
    account_product = '' # # Manual if not file pos [col, row] can be set
    validity_date_col = 4 #Column for "valid date" value 
    #validity_date_pos = [x, y] # Read from a specific file position if not on every network row
    currency_pos = 'EUR' # Manual if not file pos [col, row] can be set
    data_start_row = 19 # First row of network data
    mccmnc_col = 3 # Column for concatenated mccmnc value 
    rate_col = 5 # Column for concatenated rate value 
    change_status_col = 6 # Column "Increase", "Decrease", "Unchanged" value
    fx_conversion = 100 # Conversion multiplication value for FX is "1", "100" 
                        # or "0.01". I.e if supplier rate is in EUR it has to be 
                        # mulitplied by 100 to become EURO cent as in system.
    rate_decimal_type = ',' # If ',' (comma), replace to '.' (full stop) as system file 
                            # requires it. IGNORE coding it for now, seems Python takes care of it.                    

    # Define dict that will hold all new MCCMNC-RATE value pairs
    new_rates = {} 
   
    # Open supplier excel file and add key-value pairs to dict
    try:
        wb = openpyxl.load_workbook(new_cost_file, data_only=True)
        sheet = wb['Price List']
        print('Sheet max rows: %s and file path %s' % (sheet.max_row, new_cost_file))
        cnt = 0
        for rowNum in range(data_start_row, sheet.max_row + 1): # skip headers, data starts on row 6
            mccmnc = sheet.cell(row=rowNum, column=mccmnc_col).value # Set column for mccmnc
            current_rate = sheet.cell(row=rowNum, column=rate_col).value # Set column for current_rate
            validity_date = sheet.cell(row=rowNum, column=validity_date_col).value # Assuming its the same and given (not blank cell) for all networks
            change_status = sheet.cell(row=rowNum, column=change_status_col).value # Unique per network
            new_rates[mccmnc] = current_rate * fx_conversion # store a rate to a network in the dict
            #new_rates[mccmnc] = str(int(current_rate) * fx_conversion) # Convert to Euro cent (requires dot instead of comma for rate)
            cnt += 1
        wb.close()
        
       # pprint.pprint(new_rates)
        print('Supplier:', supplier_name_pos)
        print ('Max rows in file: ', sheet.max_row)
        print('Count of rows iterated:', cnt)
        print('Network list lenght:', str(len(new_rates)))
        print('Currency in file:', currency_pos)
        print('Validity date:', validity_date)
       
        return new_rates
     
    except Exception as e: print('An error occured in function create_dict_from_new(): ' + str(e)) 

def create_dict_from_system(system_current):
    """ Create a dict with current rates derived from the downloaded system ratedeck of supplier in CSV format.
    ARGS: system current ratedeck in CSV format
    RETURN: dict with MCCMNC-RATE value pairs of new pricing
    """
    ## Dev status: 
    # 2019-11-14 Done and tested!

    # Open and read current system CSV file
    try:
        rates_current_dict={} # dict that will hold all current MCCMNC prices for supplier
        print('Loading current_file from system...')
        with open(system_current) as csvfile:
            reader = csv.reader(csvfile, delimiter=';')
            ratelist = list(reader)
            print('Creating current_rate dictionary from current_file...')
            cnt = 0
            for i in range(1, len(ratelist)): # skip header row
                mccmnc = ratelist[i][6] + ratelist[i][7] 
                (key, val) = mccmnc, ratelist[i][8] 
                rates_current_dict[str(key)] = val
                cnt +=1
        #pprint.pprint(rates_current_dict)
        print('Number of current networks processed: ', cnt)
        print ('Number of current networks in list: ', str(len(rates_current_dict)))
        print ('Number of networks with MCC = 0 and MNC = na and default rate = 2.0 EURC: ', str(cnt-len(rates_current_dict)))
        return rates_current_dict

    except Exception as e: print('An error occured in function create_dict_from_system(): ' + str(e))   

def create_new_system_file(system_current, supplier_new, outputfile):
    """
    Takes an EBSS supplier rate deck CSV file and updates all new rates (only) based on MCCMNC identifier from new_rates dict
    created by function create_dict_from_new()
    PARAM: a downloaded suppliler rate deck from system, supplier_new file (dict creatation will be made by funcation call inside)
    RETURN: a system uploadble rate deck with new_rates added (only new rates will be updated and only for existing MCCMNC)
    """
    ## Dev status: 
    # 2019-11-14 Tested and works. But how to handle added reach in new_file that doesn't exist in current_file?
    # For example 41288 is added in new rate deck but doesn't exist in current_file (or system as network id).
    # Answer. Ignore networks not in the current_file. Write the added network to log for later implementaion 
    # in system. 
    
    # Open current_file for read and create a new_file for write
    try:
        # open input file and add each data row to a list
        print('Reading current system file...')
        with open(system_current) as csvfile:
            reader = csv.reader(csvfile, delimiter=';')
            ratelist = list(reader)

        # Call function to get new_rates dict 
        print ('Fethcing new rates dict from new_file...')
        new_rates =  create_dict_from_supplier_file(supplier_new)               
          
        # Create result file for read
        resultFile = open(outputfile, 'w')
        print('Updating current_file with new rates...')
        # Loop through and write new rates to file
        for i in range(0, len(ratelist) ):
            mccmnc = ratelist[i][6] + ratelist[i][7]  # join mcc and mnc for dict lookup
            if mccmnc in new_rates:
                ratelist[i][8] = new_rates[mccmnc]
            
            # Write data to new file, make it semi colon separated
            resultFile.write(
                ratelist[i][0] + ';' + # region
                ratelist[i][1] + ';' + # country
                str(ratelist[i][2]) + ';' + # op id
                ratelist[i][3] + ';' + # operator
                ratelist[i][4] + ';' + # network
                ratelist[i][5] + ';' + # unique name
                str(ratelist[i][6]) + ';' + # MCC
                str(ratelist[i][7]) + ';' + # MNC
                str(ratelist[i][8]) + '\n') # rate  
           
        print('Total networks processed:', i)
        print('Number of network rates updated updated: ', str(len(new_rates)))
        print('Path to new file system file for upload: ', outputfile)
        resultFile.close() 
        print('Operation complete!')        
    except Exception as e: print('An error occured in function create_new_system_file(): ' + str(e))                      

def compare_dicts(dict_new, dict_current):
    """ Takes two (network-rate key value pairs) dicts and compares them. Typically this function would be used to compare a new
    rate deck with an old. It's no point comparing a new rate deck with a downloaded system file since the system
    file will hold a lot inactive networks. 
    WHAT: Compare the current and new file rate changes and  highlights differences, ie added or removed networks
    and rate changes for the existing networks. 
    ARGS: two dicts with mccmnc:rate key value pairs
    RETURNS: delta file or print out
    
     """
    ## Dev status: 
    # 2019-11-14 Done and tested!
    # Rand values as not error handled giving for example Set()

    try:
        print('Number of networks in dict_new:', len(dict_new))
        print('Number of networks in dict_current:', len(dict_current))
        
        ## WHAT: compare two MCCMNC:rate dicts with mccmnc key as common denominator
        # price change for each MCCMNC (not implemented)
        # Lost MCCMNC's
        # Added MCCMNC's

        # Compare lists and for each key (mccmnc) display the new value and in second column
        # What kind of change type (increase/decreas/added/removed/same)
        # Result = {mccmnc:[new_value, change_type]}

        # Create a dict holding the updates (to start with)
        # Check what diffs between dicts add to diff_item dict
        diff_items = {k: dict_new[k] for k in dict_new if k in dict_current and dict_new[k] != dict_current[k]}
        print('Number of networks with differences when both files compared:', len(diff_items))
        print(' \n List with networks with differences and their latest values from new file: ')
        pprint.pprint(diff_items, width=1)
        #print('New dict:')
        #pprint.pprint(dict_new)
        #print('Current dict:')
        #pprint.pprint(dict_current)

        print(' \n Unchanged networks in new compared to current: ')
        pprint.pprint(dict_current.items() & dict_new.items())
        print(' \n Number of networks with changes in new compared to current: ', len(dict_new) - len(dict_current.items() & dict_new.items()))

        # When comparing diffs the smaller dict needs to be subtracted from the larger dict...which is why below check is needed
        if len(dict_current) > len(dict_new):
            print('Find difference in keys, current - new, i.e. lost reach, unique in current and missing in new:')
            pprint.pprint(dict_current.keys() - dict_new.keys())
            # create dict with updated newtworks
            #update_dict =  diff_items & (dict_current.keys() - dict_new.keys())
        else:
            print('Find difference in keys, new - current, i.e. added reach, unique networks in new list and missing in current:')
            pprint.pprint(dict_new.keys() - dict_current.keys())
            # create dict with updated newtworks
            #update_dict =  diff_items & (dict_new.keys() - dict_current.keys())


    except Exception as e: print('An error occured in function compare_dicts(): ' + str(e))  

def compare_rate_files(new_cost_file, old_cost_file):
    
    """
     WHAT: Compares two ratedeck files, as is, from supplier in its unique format anc column configuration e.g xlsx
     WHY: Compare two ratedecks to highlight updates
     ARGS: Two supplier files, new and old
     RETURN: added / lost reach, rate changes per network (increase, dcrease, unchanged)

    """   
    network_rates_new = {} 
    network_rates_old = {} 

    try:

        # Read the new file
        wb_new = openpyxl.load_workbook(new_cost_file, data_only=True)
        sheet = wb_new['Price List']
        print ('New_file -------------------------')
        print('Sheet max rows: %s and new file path %s' % (sheet.max_row, new_cost_file))
        cnt = 0
        for rowNum in range(data_start_row, sheet.max_row + 1): # skip headers, data starts on row 6
            mccmnc = sheet.cell(row=rowNum, column=mccmnc_col).value # Set column for mccmnc
            current_rate = sheet.cell(row=rowNum, column=rate_col).value # Set column for current_rate
            validity_date = sheet.cell(row=rowNum, column=validity_date_col).value # Assuming its the same and given (not blank cell) for all networks
            change_status = sheet.cell(row=rowNum, column=change_status_col).value # Unique per network
            network_rates_new[mccmnc] = current_rate * fx_conversion # store a rate to a network in the dict
            #new_rates[mccmnc] = str(int(current_rate) * fx_conversion) # Convert to Euro cent (requires dot instead of comma for rate)
            cnt += 1
        wb_new.close()
        
        #pprint.pprint(network_rates_new)
        print('Supplier:', supplier_name_pos)
        print ('Max rows in file: ', sheet.max_row)
        print('Count of rows iterated:', cnt)
        print('Network list lenght:', str(len(network_rates_new )))
        print('Currency in file:', currency_pos)
        print('Validity date:', validity_date)
        print('Change status:', change_status)
       

        # Read the old file

        wb_old = openpyxl.load_workbook(old_cost_file, data_only=True)
        sheet = wb_old['Price List']
        print ('Old_file -------------------------')
        print('Sheet max rows: %s and old file path %s' % (sheet.max_row, old_cost_file))
        cnt = 0
        for rowNum in range(data_start_row, sheet.max_row + 1): # skip headers, data starts on row 6
            mccmnc = sheet.cell(row=rowNum, column=mccmnc_col).value # Set column for mccmnc
            current_rate = sheet.cell(row=rowNum, column=rate_col).value # Set column for current_rate
            validity_date = sheet.cell(row=rowNum, column=validity_date_col).value # Assuming its the same and given (not blank cell) for all networks
            change_status = sheet.cell(row=rowNum, column=change_status_col).value # Unique per network
            network_rates_old[mccmnc] = current_rate * fx_conversion # store a rate to a network in the dict
            #new_rates[mccmnc] = str(int(current_rate) * fx_conversion) # Convert to Euro cent (requires dot instead of comma for rate)
            cnt += 1
        wb_old.close()
        
       
        #pprint.pprint(network_rates_old)
        print('Supplier:', supplier_name_pos)
        print ('Max rows in file: ', sheet.max_row)
        print('Count of rows iterated:', cnt)
        print('Network list lenght:', str(len(network_rates_new )))
        print('Currency in file:', currency_pos)
        print('Validity date:', validity_date)
        print('Change status:', change_status)
        
        #return network_rates_new, network_rates_old

        ## Call compare function
        print(' \nNew vs old comparison --------------------')
        
        # call compare function
        compare_dicts(network_rates_new, network_rates_old)

    except Exception as e: print('An error occured in function compare_rate_files(): ' + str(e))   

#### MAIN #####

# File logistics:

## Downloaded from system to drive folder
system_current = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\generic_supplier\\EBSS\\2019-11-18 price_list_DEUT.csv"
## New file received from supplier
supplier_new = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\\\generic_supplier\\Deutsche Telekom AG\\2019-11-12 priceList [dtagtele2554][A-Z].xlsx"

# Path to result file for upload back to system
outputfile = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\generic_supplier\\Results\\new_system_file.csv"

## Create new rate dictionary
#create_dict_from_new(supplier_new)

## Create current rate dictionary
#create_dict_from_system(system_current)

## Create a new system file
#create_new_system_file(system_current, supplier_new, outputfile)

## Compare files
#new_cost_file = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\\\generic_supplier\\Deutsche Telekom AG\\2019-11-18 priceList [dtagtele2554][A-Z].xlsx"
new_cost_file = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\\\generic_supplier\\Deutsche Telekom AG\\2019-11-12 priceList [dtagtele2554][A-Z].xlsx"
old_cost_file = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\\\generic_supplier\\Deutsche Telekom AG\\2019-11-11 priceList [dtagtele2554][A-Z].xlsx"

compare_rate_files(new_cost_file, old_cost_file)
