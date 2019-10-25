import openpyxl, sys
from openpyxl import load_workbook

def splitMNC():
    # C:/Users/johastje/AppData/Local/Programs/Python/Python36/python.exe C:\Users\johastje\Desktop\PythonPlay\Costlists\split_MNC\splitMNC.py
    """
    Splits MNC's that are grouped in one cell with all same price
    PARAM: input cost file with destinations and costs
    RETURN: output file
    """
    # 1.Open the spreadsheet file.
    # 2.For each row, check whether the value in MNC column has grouped MNC's in a cell
    # 3.If it is, move each MNC on its own row and copy all other column dara
    # 4.Save the spreadsheet to a new file (so that you don’t lose the old spreadsheet, just in case).

    inputfile = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\split_MNC\\split_mnc_input_data.xlsx"
    outputfile = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Costlists\\split_MNC\\splitMNCdata_result.csv" 
    try:
        print('Reading input file...')
        wb = openpyxl.load_workbook(inputfile)
        sheet = wb['sheet1']
        resultFile = open(outputfile, 'w')

        # TODO: Loop through the rows and check for separators
        # Store MNC to an array for later writing to new file
        # If separator in loop through each mnc and print a row for each
        print('Parsing and writing to new file...')
        rowcnt = 0
        mnccnt = 0
        for rowNum in range(2, sheet.max_row+1): # skip until data row
            rowcnt += 1
            mnc = sheet.cell(row=rowNum, column=3).value
            mnc = mnc.replace(" ","") # clean white space, mnc.strip()
            numMNC = mnc.count(';') + 1 # count number of mnc's
            # List the column values of interest
            country = sheet.cell(row=rowNum, column=1).value
            mcc = sheet.cell(row=rowNum, column=2).value
            rate = sheet.cell(row=rowNum, column=4).value
            # if more than one mnc in cell, put each in arr split on separator
            if numMNC > 1:
                arr = mnc.split(';')
                for item in arr:
                    # write rest of the row columns
                    resultFile.write(country + ';' + str(mcc) + ';' + str(item) + ';' + str(rate) + '\n')
                    mnccnt +=1 
            # write to file and keep looping
            else:
                resultFile.write(country + ';' + str(mcc) + ';' + str(mnc) + ';' + str(rate) + '\n')
        resultFile.close()
        print('Rows processed:', rowcnt)
        print('MNCs processed:', mnccnt)
        print('Operation completed!')   
    except Exception as e: 
        print('An error occured: ' + str(e))

def convertExelToCSV():  
    """ Reads an Excel file and converts to a CSV file, """

    in_path = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Pricelists\\test.xlsx" 
    try:
        wb = openpyxl.load_workbook(in_path)
        sheet = wb['Sheet1']

        # Initiate csv writer
        out_path = "C:\\Users\\johastje\\Desktop\\PythonPlay\\Pricelists\\output1.csv"
        outputFile = open(out_path, 'w', newline='')
        outputWriter = csv.writer(outputFile)

        # Loop print each cell value for a row to list
        rowData = []
        row = 0
        for rowNum in range( 1, sheet.max_row + 1):
            for colNum in range( 1, sheet.max_column + 1):
                row = row + sheet.cell(row=rowNum, column=colNum).value
            # rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                if colNum == sheet.max_column:
                    rowData.append("\n")

        # print list to file        
        outputWriter.writerow(rowData)
        outputFile.close() 
        print(rowData)
        print("File converted!")  

    except Exception as e: print('An error occured in convertToCSV: ' + str(e))     
# main
splitMNC()         